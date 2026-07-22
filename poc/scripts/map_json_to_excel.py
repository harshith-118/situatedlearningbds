#!/usr/bin/env python3
"""PoC: map AI extraction JSON into Excel template rows.

This script simulates the post-AI code-execution step in the Situated Learning
pipeline. It does NOT call any real client systems and uses only synthetic JSON.
"""
from __future__ import annotations

import json
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1] / "sample_data"
JSON_DIR = ROOT / "extracted" / "json"
OUT = ROOT / "exports" / "Contract_Extract_Filled_PoC.xlsx"

HEADERS = [
    "Contract ID", "Client Name", "Effective Date", "Termination Date",
    "Plan Type", "Network Name", "Coverage Limit (USD)", "Deductible (USD)",
    "Co-pay (USD)", "Region", "Broker", "Status", "Source Format",
    "Extraction Status", "Missing Fields", "Run ID",
]

def main() -> None:
    payloads = []
    for path in sorted(JSON_DIR.glob("*.json")):
        payloads.append(json.loads(path.read_text()))

    wb = Workbook()
    ws = wb.active
    ws.title = "Contract Extract"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 16

    for r, p in enumerate(payloads, 2):
        f = p["fields"]
        vals = [
            f.get("contract_id"), f.get("client_name"), f.get("effective_date"),
            f.get("termination_date"), f.get("plan_type"), f.get("network_name"),
            f.get("coverage_limit"), f.get("deductible"), f.get("copay"),
            f.get("region"), f.get("broker"), f.get("status"),
            p.get("source_format"), p.get("extraction_status"),
            ",".join(p.get("missing_fields") or []), p.get("run_id"),
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(r, c, "" if v is None else v)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT} with {len(payloads)} rows")

if __name__ == "__main__":
    main()
